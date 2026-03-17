"""
Parse ALL ITEMS.xlsx and generate SQL to seed the stock management system.
Maps: Warehouse = Branch (KZN, GP, PE, CPT, PVT, BOND)
      Building = Division (PROMED, ACCESS, SEB, KOG)
      BuildingInventory = Items with quantities per division per branch

Imports ALL 794 items into all relevant buildings (even with 0 qty)
so the full product catalog is visible in every location.
"""
import openpyxl
import os

# Column mapping (1-indexed) from the Excel
# Division -> { branch: column_number }
DIVISION_BRANCH_COLS = {
    'ProMed': {
        'KZN': 4, 'GP': 5, 'PE': 6, 'CPT': 7, 'PVT': 8
    },
    'Access': {
        'KZN': 9, 'GP': 10, 'PE': 11, 'CPT': 12, 'PVT': 13
    },
    'SEB': {
        'KZN': 14, 'GP': 15, 'PE': 16, 'CPT': 17
    },
    'KOG': {
        'KZN': 18, 'GP': 19, 'PE': 20, 'CPT': 21, 'BOND': 22
    }
}

# Warehouse definitions
WAREHOUSES = {
    'KZN':  {'id': 1, 'name': 'KZN Warehouse',  'code': 'WH-KZN',  'city': 'Durban',    'province': 'KwaZulu-Natal'},
    'GP':   {'id': 2, 'name': 'GP Warehouse',   'code': 'WH-GP',   'city': 'Johannesburg','province': 'Gauteng'},
    'CPT':  {'id': 3, 'name': 'CPT Warehouse',  'code': 'WH-CPT',  'city': 'Cape Town',  'province': 'Western Cape'},
    'PE':   {'id': 4, 'name': 'PE Warehouse',   'code': 'WH-PE',   'city': 'Gqeberha',  'province': 'Eastern Cape'},
    'PVT':  {'id': 5, 'name': 'PVT Warehouse',  'code': 'WH-PVT',  'city': 'Private',    'province': 'National'},
    'BOND': {'id': 6, 'name': 'BOND Warehouse', 'code': 'WH-BOND', 'city': 'Bond Store', 'province': 'National'},
}


def get_building_id(warehouse_id, division_name):
    div_idx = {'ProMed': 1, 'Access': 2, 'SEB': 3, 'KOG': 4}
    return warehouse_id * 10 + div_idx[division_name]


def escape_sql(s):
    if s is None:
        return 'NULL'
    return "'" + str(s).replace("'", "''").strip() + "'"


def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    items = []
    for row_idx in range(4, ws.max_row + 1):
        code = ws.cell(row=row_idx, column=1).value
        desc = ws.cell(row=row_idx, column=2).value
        uom = ws.cell(row=row_idx, column=3).value
        
        if code is None and desc is None:
            continue
        
        item = {
            'code': str(code).strip() if code else '',
            'description': str(desc).strip() if desc else '',
            'uom': str(uom).strip() if uom else 'Each',
            'quantities': {}  # {(division, branch): qty}
        }
        
        # Parse system quantities per division per branch
        for division, branches in DIVISION_BRANCH_COLS.items():
            for branch, col in branches.items():
                val = ws.cell(row=row_idx, column=col).value
                try:
                    qty = float(val) if val is not None else 0
                except (ValueError, TypeError):
                    qty = 0
                item['quantities'][(division, branch)] = qty
        
        if item['code'] or item['description']:
            items.append(item)
    
    return items


def generate_sql(items):
    sql_lines = []
    sql_lines.append("-- ============================================================")
    sql_lines.append("-- Stock Management Seed Data from ALL ITEMS.xlsx")
    sql_lines.append("-- Generated automatically")
    sql_lines.append("-- Maps: Warehouse=Branch, Building=Division, BuildingInventory=Items")
    sql_lines.append("-- ============================================================")
    sql_lines.append("")
    sql_lines.append("USE ProjectTrackerDB;")
    sql_lines.append("GO")
    sql_lines.append("")
    
    # ---- Step 1: Update existing warehouses ----
    sql_lines.append("-- Step 1: Update existing warehouses to match branch codes")
    sql_lines.append("PRINT 'Updating existing warehouses...';")
    sql_lines.append("")
    sql_lines.append("UPDATE Warehouses SET Code = 'WH-CPT', Name = 'CPT Warehouse' WHERE Id = 3;")
    sql_lines.append("UPDATE Warehouses SET Code = 'WH-PE', Name = 'PE Warehouse' WHERE Id = 4;")
    sql_lines.append("")
    
    # ---- Step 2: Insert new warehouses ----
    sql_lines.append("-- Step 2: Add PVT and BOND warehouses")
    sql_lines.append("PRINT 'Adding PVT and BOND warehouses...';")
    sql_lines.append("")
    sql_lines.append("SET IDENTITY_INSERT Warehouses ON;")
    sql_lines.append("")
    
    for branch_code in ['PVT', 'BOND']:
        wh = WAREHOUSES[branch_code]
        sql_lines.append(f"IF NOT EXISTS (SELECT 1 FROM Warehouses WHERE Id = {wh['id']})")
        sql_lines.append(f"INSERT INTO Warehouses (Id, Name, Code, City, Province, TotalCapacity, AvailableCapacity, Status, CreatedAt)")
        sql_lines.append(f"VALUES ({wh['id']}, {escape_sql(wh['name'])}, {escape_sql(wh['code'])}, {escape_sql(wh['city'])}, {escape_sql(wh['province'])}, 50000, 50000, 'Active', GETUTCDATE());")
        sql_lines.append("")
    
    sql_lines.append("SET IDENTITY_INSERT Warehouses OFF;")
    sql_lines.append("")
    
    # ---- Step 3: Create buildings (divisions) ----
    sql_lines.append("-- Step 3: Create division buildings under each warehouse")
    sql_lines.append("PRINT 'Creating division buildings...';")
    sql_lines.append("")
    sql_lines.append("SET IDENTITY_INSERT WarehouseBuildings ON;")
    sql_lines.append("")
    
    buildings_created = {}
    
    for division, branches in DIVISION_BRANCH_COLS.items():
        for branch in branches.keys():
            wh = WAREHOUSES[branch]
            building_id = get_building_id(wh['id'], division)
            building_code = f"{branch}-{division.upper()}"
            building_name = division
            
            buildings_created[(branch, division)] = building_id
            
            sql_lines.append(f"IF NOT EXISTS (SELECT 1 FROM WarehouseBuildings WHERE Id = {building_id})")
            sql_lines.append(f"INSERT INTO WarehouseBuildings (Id, WarehouseId, Code, Name, IsActive, CreatedAt, UpdatedAt)")
            sql_lines.append(f"VALUES ({building_id}, {wh['id']}, {escape_sql(building_code)}, {escape_sql(building_name)}, 1, GETUTCDATE(), GETUTCDATE());")
            sql_lines.append("")
    
    sql_lines.append("SET IDENTITY_INSERT WarehouseBuildings OFF;")
    sql_lines.append("")
    
    # ---- Step 4: Insert BuildingInventory ----
    sql_lines.append("-- Step 4: Import ALL items as BuildingInventory")
    sql_lines.append(f"PRINT 'Importing {len(items)} items into BuildingInventory...';")
    sql_lines.append("")
    sql_lines.append("-- Clear any existing building inventory")
    sql_lines.append("DELETE FROM BuildingInventory;")
    sql_lines.append("")
    
    insert_count = 0
    nonzero_count = 0
    batch_size = 50
    batch = []
    
    for item in items:
        item_code = item['code'][:50] if item['code'] else f'ITEM-{insert_count:05d}'
        item_desc = item['description'][:500] if item['description'] else item_code
        uom = item['uom'][:20] if item['uom'] else 'Each'
        
        for (division, branch), qty in item['quantities'].items():
            if (branch, division) not in buildings_created:
                continue
            
            building_id = buildings_created[(branch, division)]
            
            batch.append(
                f"({building_id}, {escape_sql(item_code)}, {escape_sql(item_desc)}, {escape_sql(uom)}, "
                f"{qty}, 0, 0, NULL, NULL, NULL, NULL, NULL, NULL, GETUTCDATE(), GETUTCDATE())"
            )
            insert_count += 1
            if qty > 0:
                nonzero_count += 1
            
            if len(batch) >= batch_size:
                sql_lines.append("INSERT INTO BuildingInventory (BuildingId, ItemCode, ItemDescription, Uom, QuantityOnHand, QuantityReserved, QuantityOnOrder, ReorderLevel, MaxLevel, UnitCost, BinLocation, LastMovementDate, LastCountDate, CreatedAt, UpdatedAt)")
                sql_lines.append("VALUES")
                sql_lines.append(",\n".join(batch) + ";")
                sql_lines.append("")
                batch = []
    
    if batch:
        sql_lines.append("INSERT INTO BuildingInventory (BuildingId, ItemCode, ItemDescription, Uom, QuantityOnHand, QuantityReserved, QuantityOnOrder, ReorderLevel, MaxLevel, UnitCost, BinLocation, LastMovementDate, LastCountDate, CreatedAt, UpdatedAt)")
        sql_lines.append("VALUES")
        sql_lines.append(",\n".join(batch) + ";")
        sql_lines.append("")
    
    # ---- Step 5: Summary ----
    sql_lines.append("-- Step 5: Verify results")
    sql_lines.append("PRINT 'Seed complete. Verifying...';")
    sql_lines.append("")
    sql_lines.append("SELECT 'Warehouses' AS [Table], COUNT(*) AS [Count] FROM Warehouses;")
    sql_lines.append("SELECT 'Buildings' AS [Table], COUNT(*) AS [Count] FROM WarehouseBuildings;")
    sql_lines.append("SELECT 'Inventory Items' AS [Table], COUNT(*) AS [Count] FROM BuildingInventory;")
    sql_lines.append("SELECT 'Items with Stock' AS [Metric], COUNT(*) AS [Count] FROM BuildingInventory WHERE QuantityOnHand > 0;")
    sql_lines.append("")
    sql_lines.append("SELECT w.Name AS Warehouse, wb.Name AS Division, COUNT(bi.Id) AS Items,")
    sql_lines.append("  SUM(CASE WHEN bi.QuantityOnHand > 0 THEN 1 ELSE 0 END) AS WithStock,")
    sql_lines.append("  CAST(SUM(bi.QuantityOnHand) AS INT) AS TotalQty")
    sql_lines.append("FROM Warehouses w")
    sql_lines.append("JOIN WarehouseBuildings wb ON wb.WarehouseId = w.Id")
    sql_lines.append("LEFT JOIN BuildingInventory bi ON bi.BuildingId = wb.Id")
    sql_lines.append("GROUP BY w.Name, wb.Name")
    sql_lines.append("ORDER BY w.Name, wb.Name;")
    sql_lines.append("")
    sql_lines.append(f"PRINT 'Successfully imported {insert_count} inventory records ({nonzero_count} with stock) from {len(items)} items.';")
    sql_lines.append("GO")
    
    return "\n".join(sql_lines), insert_count, nonzero_count


if __name__ == '__main__':
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(os.path.dirname(script_dir), 'ALL ITEMS.xlsx')
    
    print(f"Reading Excel from: {excel_path}")
    items = parse_excel(excel_path)
    print(f"Parsed {len(items)} items from Excel")
    
    sql, total, with_stock = generate_sql(items)
    
    output_path = os.path.join(script_dir, 'SeedStockFromExcel.sql')
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(sql)
    
    print(f"Generated SQL: {total} inventory records ({with_stock} with stock) -> {output_path}")
