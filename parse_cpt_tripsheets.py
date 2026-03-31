import openpyxl, os, re, json

folder = r'C:\Users\Administrator\Desktop\CPT TRIPSHEET 2026\CPT TRIPSHEET 2026'
all_data = {}

for fname in sorted(os.listdir(folder)):
    if not fname.endswith('.xlsx'):
        continue
    path = os.path.join(folder, fname)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Sheet1']
    
    # Get driver from C2, date from J2
    driver = str(ws['C2'].value or '').strip()
    date_val = str(ws['J2'].value or '').strip()
    
    # Get company indicator from B column
    invoices = []
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=False):
        inv = row[0].value  # Column A = INV NO.
        if inv and str(inv).strip():
            inv_str = str(inv).strip()
            # Only keep actual invoice numbers (start with IN)
            if re.match(r'^IN\d+', inv_str):
                company_no = str(row[1].value or '').strip()  # Column B = No. (company indicator)
                cust = str(row[2].value or '').strip()
                product = str(row[3].value or '').strip()
                brand = str(row[4].value or '').strip()
                qty = str(row[5].value or '').strip()
                time_disp = str(row[6].value or '').strip()
                order_no = str(row[7].value or '').strip()
                val = row[11].value  # Column L = VALUE
                invoices.append({
                    'inv': inv_str,
                    'company': company_no,
                    'cust': cust,
                    'product': product,
                    'brand': brand,
                    'qty': qty,
                    'order': order_no,
                    'val': val
                })
    
    all_data[fname] = {
        'driver': driver,
        'date': date_val,
        'invoices': invoices,
        'count': len(invoices)
    }
    print(f"{fname}: Driver={driver}, Date={date_val}, Invoices={len(invoices)}")
    for inv in invoices:
        print(f"  {inv['inv']} | {inv['cust'][:60]} | Val={inv['val']}")

print(f"\nTotal files: {len(all_data)}")
total_inv = sum(d['count'] for d in all_data.values())
print(f"Total invoices across all files: {total_inv}")

# Collect all unique invoice numbers
all_invs = set()
for d in all_data.values():
    for inv in d['invoices']:
        all_invs.add(inv['inv'])
print(f"Unique invoice numbers: {len(all_invs)}")

# Save for later use
with open('cpt_tripsheet_data.json', 'w') as f:
    json.dump(all_data, f, indent=2, default=str)
print("\nSaved to cpt_tripsheet_data.json")
