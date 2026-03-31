import openpyxl, os, re, json

folder = r'docs\KZN TRIPSHEET'
all_data = {}
skipped = []

for root, dirs, files in os.walk(folder):
    for fname in sorted(files):
        if not fname.endswith('.xlsx') or fname.startswith('~'):
            continue
        path = os.path.join(root, fname)
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active

            # Get driver from C2 (clean trailing whitespace/newlines)
            driver = str(ws['C2'].value or '').strip().replace('\n', '').replace('\r', '').upper()

            # Date: try J2 first, then L2
            date_val = str(ws['J2'].value or '').strip()
            if not date_val or date_val == 'None':
                date_val = str(ws['L2'].value or '').strip()

            # Vehicle registration from C3
            vehicle_reg = str(ws['C3'].value or '').strip().upper()

            # Route name from filename (strip date and extension)
            route = re.sub(r'\s*\d{2}\.\d{2}\.\d{2,4}.*$', '', fname.replace('.xlsx', '')).strip()
            route = re.sub(r'\s*-\s*Copy$', '', route).strip()
            route = re.sub(r'^Copy of\s*', '', route).strip()

            # Extract invoices from row 6+
            invoices = []
            for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=False):
                inv = row[0].value  # Column A = INV NO.
                if inv and str(inv).strip():
                    inv_str = str(inv).strip()
                    if re.match(r'^IN\d+', inv_str):
                        company_no = str(row[1].value or '').strip()
                        cust = str(row[2].value or '').strip()
                        product = str(row[3].value or '').strip()
                        brand = str(row[4].value or '').strip()
                        qty = str(row[5].value or '').strip()
                        time_disp = str(row[6].value or '').strip()
                        order_no = str(row[7].value or '').strip()
                        val = row[11].value if len(row) > 11 else None  # Column L = VALUE
                        invoices.append({
                            'inv': inv_str,
                            'company': company_no,
                            'cust': cust,
                            'product': product,
                            'brand': brand,
                            'qty': qty,
                            'order': order_no,
                            'val': val
                        })

            # Use relative path as key to avoid collisions between months
            rel_path = os.path.relpath(path, folder).replace('\\', '/')
            all_data[rel_path] = {
                'filename': fname,
                'driver': driver,
                'date': date_val,
                'vehicle_reg': vehicle_reg,
                'route': route,
                'invoices': invoices,
                'count': len(invoices)
            }
            print(f"{rel_path}: Driver={driver or '(none)'}, Date={date_val}, Route={route}, Invoices={len(invoices)}")

        except Exception as e:
            skipped.append(f"{fname}: {e}")
            print(f"ERROR {fname}: {e}")

print(f"\nTotal files parsed: {len(all_data)}")
total_inv = sum(d['count'] for d in all_data.values())
print(f"Total invoices across all files: {total_inv}")

# Unique invoice numbers
all_invs = set()
for d in all_data.values():
    for inv in d['invoices']:
        all_invs.add(inv['inv'])
print(f"Unique invoice numbers: {len(all_invs)}")

# Unique drivers
drivers = sorted(set(d['driver'] for d in all_data.values()))
print(f"Unique drivers: {drivers}")

# Unique routes
routes = sorted(set(d['route'] for d in all_data.values()))
print(f"Unique routes ({len(routes)}): {routes}")

if skipped:
    print(f"\nSkipped {len(skipped)} files:")
    for s in skipped:
        print(f"  {s}")

with open('kzn_tripsheet_data.json', 'w') as f:
    json.dump(all_data, f, indent=2, default=str)
print("\nSaved to kzn_tripsheet_data.json")
